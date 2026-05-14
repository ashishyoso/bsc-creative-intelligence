"""Parse the Hawky XLSX/CSV export into normalized rows.

Implements US-1.2:
- Multi-row Ad IDs split into separate Ad IDs all linked to one creative
- Percentages stored as decimals
- ROAS precision is preserved; integer-rounded ROAS flagged
- Missing cells preserved as None (distinguish 'no data' from 'zero')
"""
from __future__ import annotations

import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

import openpyxl


HAWKY_COLUMNS = {
    "Creative": "cdn_url",
    "Ad ID": "ad_ids",
    "Ad Name": "ad_names",
    "Spend": "spend",
    "ROAS": "roas",
    "Impressions": "impressions",
    "Hook Rate": "hook_rate",
    "Hold Rate": "hold_rate",
    "CTR": "ctr",
    "Clicks": "clicks",
    "Add To Cart": "atc",
    "All Clicks": "all_clicks",
    "CPC": "cpc",
    "CPM": "cpm",
    "Creative Preview": "creative_preview",
}


@dataclass
class HawkyRow:
    row_index: int
    cdn_url: str | None
    ad_ids: list[str] = field(default_factory=list)
    ad_names: list[str] = field(default_factory=list)
    primary_ad_id: str | None = None
    primary_ad_name: str | None = None
    spend: float | None = None
    roas: float | None = None
    impressions: int | None = None
    hook_rate: float | None = None
    hold_rate: float | None = None
    ctr: float | None = None
    clicks: int | None = None
    atc: int | None = None
    all_clicks: int | None = None
    cpc: float | None = None
    cpm: float | None = None
    creative_preview: str | None = None
    raw_roas_was_integer: bool = False


def _split_multi(value: Any) -> list[str]:
    """Split comma/newline-separated Ad IDs or Ad Names. Trim aggressively."""
    if value is None:
        return []
    if not isinstance(value, str):
        value = str(value)
    parts = re.split(r"[,\n]+", value)
    return [p.strip() for p in parts if p and p.strip()]


def _to_float(value: Any) -> float | None:
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        return float(value)
    s = str(value).strip().rstrip("%")
    if not s:
        return None
    try:
        return float(s)
    except ValueError:
        return None


def _to_int(value: Any) -> int | None:
    f = _to_float(value)
    if f is None:
        return None
    return int(f)


def _percent_to_decimal(value: Any) -> float | None:
    """Hawky's hook/hold/CTR may be either fractions (0.21) or percentages (21.85).
    We normalize to decimals: if value > 1.5 we assume it's a percent."""
    f = _to_float(value)
    if f is None:
        return None
    return f / 100.0 if f > 1.5 else f


def _classify_asset_type(cdn_url: str | None) -> str:
    if not cdn_url:
        return "video"
    url = cdn_url.lower()
    if "/images/" in url or url.endswith((".webp", ".png", ".jpg", ".jpeg", ".gif")):
        return "image"
    return "video"


def parse_hawky_xlsx(path: Path, sheet_name: str | None = None, limit: int | None = None) -> list[HawkyRow]:
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = wb[sheet_name] if sheet_name else wb[wb.sheetnames[0]]

    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    header_map = {col: idx for idx, col in enumerate(header_row)}

    missing = [col for col in HAWKY_COLUMNS if col not in header_map]
    if missing:
        raise ValueError(f"Hawky export missing expected columns: {missing}")

    rows: list[HawkyRow] = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if limit is not None and len(rows) >= limit:
            break

        def cell(name: str):
            return row[header_map[name]]

        cdn_url = cell("Creative")
        if not cdn_url:
            continue

        ad_ids = _split_multi(cell("Ad ID"))
        ad_names = _split_multi(cell("Ad Name"))
        raw_roas = cell("ROAS")
        roas = _to_float(raw_roas)
        roas_int = roas is not None and float(int(roas)) == roas

        rows.append(
            HawkyRow(
                row_index=row_idx,
                cdn_url=str(cdn_url).strip(),
                ad_ids=ad_ids,
                ad_names=ad_names,
                primary_ad_id=ad_ids[0] if ad_ids else None,
                primary_ad_name=ad_names[0] if ad_names else None,
                spend=_to_float(cell("Spend")),
                roas=roas,
                impressions=_to_int(cell("Impressions")),
                hook_rate=_percent_to_decimal(cell("Hook Rate")),
                hold_rate=_percent_to_decimal(cell("Hold Rate")),
                ctr=_percent_to_decimal(cell("CTR")),
                clicks=_to_int(cell("Clicks")),
                atc=_to_int(cell("Add To Cart")),
                all_clicks=_to_int(cell("All Clicks")),
                cpc=_to_float(cell("CPC")),
                cpm=_to_float(cell("CPM")),
                creative_preview=str(cell("Creative Preview")).strip() if cell("Creative Preview") else None,
                raw_roas_was_integer=roas_int,
            )
        )

    return rows


def detect_roas_precision_issue(rows: list[HawkyRow]) -> dict:
    """US-7.5: flag when >80% of non-zero ROAS values are integer-rounded.

    Triggers only when N>20.
    """
    nonzero = [r for r in rows if r.roas is not None and r.roas > 0]
    if len(nonzero) <= 20:
        return {"triggered": False, "reason": "n_too_small", "n": len(nonzero)}
    integer_count = sum(1 for r in nonzero if r.raw_roas_was_integer)
    pct = integer_count / len(nonzero)
    return {
        "triggered": pct > 0.80,
        "integer_pct": round(pct, 3),
        "n": len(nonzero),
        "message": (
            "Source data has integer-rounded ROAS. Pattern mining and predictions "
            "will be unreliable. Contact Hawky/Meta for 2-decimal precision."
            if pct > 0.80
            else "ROAS precision looks acceptable."
        ),
    }
